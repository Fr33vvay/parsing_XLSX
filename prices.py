def parsprice(template, month, result):
    import openpyxl

    Template = openpyxl.load_workbook(template)
    sheet_pr = Template[str(input('Введите название месяца: '))]
    Month = openpyxl.load_workbook(month)  # файл месяца
    sheet_m = Month['Лист1']

    temp_prod_list = []  # список продуктов из шаблона
    for cell in sheet_pr['A'][1:]:
        temp_prod_list.append(cell.value)

    # парсим цены из файла месяца
    month_dict = {}  # словарь продукт - цены из файла месяца
    current_product = False
    for row in sheet_m.iter_rows(min_row=2, values_only=True):
        if row[0] is not None and row[0] != 'Тип':  # убираем пустые ячейки и заголовки таблицы
            if row[0] != 'п/н':
                product = row[0].split(' (')  # отсекаем единицы измерения в названии
                if product[0] in temp_prod_list:
                    month_dict[product[0]] = []
                    current_product = product[0]  # будем добавлять цены в текущий продукт
                else:
                    current_product = False  # текущего продукта нет
            elif row[0] == 'п/н' and current_product:
                month_dict[current_product].append(row[6])  # добавляем цены

    for key in month_dict:
        month_dict[key] = [x for x in month_dict[key] if not isinstance(x, str)]  # фиксим какой-то баг со строками
        # убираем цены, когда колебания несущественны
        standard = month_dict[key][0]
        index = 1
        for element in month_dict[key][1:]:
            delta_subtr = abs(element - standard)
            delta_div = abs(element / standard)
            if (standard >=100 and delta_subtr < 10 and 0.9 < delta_div < 1.1) or (standard < 100 and delta_subtr == 0):
                del month_dict[key][index]
            else:
                standard = element
                index += 1
        if len(month_dict[key]) > 16: # сокращаем список, чтобы он влез в таблицу
            del month_dict[key][16:]

    # запись цен в новый файл
    for rownum in range(2, sheet_pr.max_row + 1):
        product_pr = sheet_pr.cell(row=rownum, column=1).value  # продукты из файла цен
        if product_pr in month_dict:
            for colnum in range(len(month_dict[product_pr])):
                sheet_pr.cell(row=rownum, column=colnum + 3).value = month_dict[product_pr][colnum]
    Template.save(result)

parsprice('Template.xlsx', 'март.xlsx', 'Результат.xlsx')